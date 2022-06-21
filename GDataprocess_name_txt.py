#!/usr/bin/python
# coding:utf-8

import base
import openpyxl
import os


"""
    產生店家名稱txt
"""


class Exc:
    oDB = base.DB()
    oDB.connect("twd_SHU")

    sProject_name = ''


    def __init__(self, iProject_no):

        aProject_data = self.oDB.select_single(
            "lists_data_project", "project_no = %s AND status = 3" % (iProject_no))
        self.sProject_name = aProject_data['project_name']

        # read Excel file
        wb = openpyxl.load_workbook(
            r"D:\\test\\Flask_test\\flask_sql\\各產業資料\\%s\\Google評分_%s_店家名單資訊.xlsx" % (self.sProject_name, self.sProject_name))

        for sheetname in wb.sheetnames:
            if sheetname == "Sheet" or sheetname == "已審核資料":
                continue

            ws = wb[sheetname]
            total_datas = []

            # Read data from second row
            for row in ws.iter_rows(min_row=2):
                line = [col.value for col in row]

                if line[4] == None:  # 僅取得未審核資料
                    total_datas.append(line)

            # 工作表內無資料者不產生檔案
            if total_datas == []:
                continue

            namedata = []

            for total_data in total_datas:
                New_txt = ""
                New_txt += str(total_data[0])
                New_txt += " >>> "
                New_txt += str(total_data[1])
                New_txt += "\n"
                namedata.append(New_txt)

            print("%s : %s 筆" % (sheetname, len(namedata)))

            # 寫入txt filename
            txt_filename = "%s_%s_list_name.txt" % (
                self.sProject_name, sheetname)

            # 寫入txt flie path
            path = r"D:\\test\\Flask_test\\flask_sql\\各產業資料\\%s\\%s" % (
                self.sProject_name, txt_filename)

            with open(path, "w", encoding="utf-8") as f:
                f.write(" 編號		店家名稱")
                f.write("\n")
                for k in namedata:
                    f.writelines(k)

        self.do_zip()

    def do_zip(self):
        """壓縮「產業」資料夾中須處理的name.txt檔案與名單excel檔

        """
        from zipfile import ZipFile, ZIP_LZMA

        sFile_path = r'D:\\test\\Flask_test\\flask_sql\\各產業資料\\%s' % (self.sProject_name)

        with ZipFile(r'%s\\%s.7z' % (sFile_path,  self.sProject_name), 'w', compression=ZIP_LZMA) as zf:

            sOrigin_path = r'%s\\origin\\' % (sFile_path)
            if not os.path.isdir(sOrigin_path):
                os.makedirs(sOrigin_path)

            for root, dirs, files in os.walk(sFile_path):
                for file in files:
                    # 僅壓縮txt file
                    if '.txt' in file:
                        zf.write(os.path.join(root, file), file)
                        print(file + ' 已加入壓縮')

                    # 指定file 壓縮
                    if (file == 'Google評分_%s_店家名單資訊.xlsx' % (self.sProject_name)):
                        zf.write(os.path.join(root, file), file)
                        print(file + ' 已加入壓縮')

            self.move_txt(sFile_path, sOrigin_path)

    def move_txt(self, sFile_path, sOrigin_path):
        """移動name_txt檔案

        Args:
            sFile_path (str): 產業資料夾
            sOrigin_path (str): 移動目標資料夾
        """
        for root, dir, files in os.walk(sFile_path):
            for file in files:
                if '.txt' in file:
                    os.rename(os.path.join(root, file), '%s\\%s' % (sOrigin_path, file))


if __name__ == '__main__':
    Exc()
