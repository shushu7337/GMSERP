# !usr/bin/python
# coding:utf-8

from openpyxl.reader.excel import load_workbook
from pymysql.converters import escape_string

import base
import json
import os

"""
    根據txt檔來回寫excel檔案,並更新資料庫
"""


class Exc:
    # 資料庫連線
    oDB = base.DB()
    oDB.connect("twd_SHU")

    sSuccess_log = ''
    sFail_log = ''
    sRemove_log = ''


    def __init__(self, sProject_no):
        """
            選擇專案
            依專案取得目標資料夾
            依目標資料夾取得目標檔案, 並執行相關更新
                1. name
                2. addr
                3. remove

                根據txt檔名判斷修改項目
                取得目標 txt檔案(name, addr, remove)
                1. 讀取,分割資料,暫存
                2. 修改資料庫資料
                3. 覆寫excel資料,存檔
        """
        try:
            aProject_data = self.oDB.select_single(
                "lists_data_project", "project_no = %s" % (sProject_no))

            sProject_name = aProject_data['project_name']
            # 各產業xt,xlsx文件路徑設定
            sNameTxtPath = r'D:\\test\\Flask_test\\flask_sql\\各產業資料\\%s\\%s_list_name_整理完成.txt' % (
                sProject_name, sProject_name)
            sAddrTxtPath = r'D:\\test\\Flask_test\\flask_sql\\各產業資料\\%s\\%s_list_addr_整理完成.txt' % (
                sProject_name, sProject_name)
            sRemoveTxtPath = r'D:\\test\\Flask_test\\flask_sql\\各產業資料\\%s\\%s_list_remove.txt' % (
                sProject_name, sProject_name)
            sExcelFilePath = r'D:\\test\\Flask_test\\flask_sql\\各產業資料\\%s\\Google評分_%s_店家名單資訊.xlsx' % (
                sProject_name, sProject_name)

            # 檢查各文件路徑 & 更新相關file資料(更新excel檔案及資料庫)
            if os.path.exists(sExcelFilePath):

                if os.path.exists(sNameTxtPath):
                    aEdited_name = self.read_txt(sNameTxtPath)
                    if aEdited_name:
                        self.update_database_data(aEdited_name)
                if os.path.exists(sAddrTxtPath):
                    aEdited_addr = self.read_txt(sAddrTxtPath)  
                    if aEdited_addr:
                        self.update_database_data(aEdited_addr)
                if os.path.exists(sRemoveTxtPath):
                    aRemove_no = self.read_txt(sRemoveTxtPath)
                    if aRemove_no:
                        self.update_database_data(aRemove_no)

                # 產生新excel
                import GDataprocess_scoredata
                GDataprocess_scoredata.Exc(sProject_no)
                
            else:
                self.sFail_log = "無 Google評分_%s_店家名單資訊 檔案" % (sProject_name)
        except Exception as e:
            print("error line: %s " % e.__traceback__.tb_lineno)
            self.sFail_log = e.__traceback__.tb_lineno


    def read_txt(self, sTxtPath):
        """讀取txt檔案

        Args:
            sTxtPath (string)): 各txt檔路徑

        Returns:
            array: 依檔案內容回傳資料
        """
        try:
            with open(sTxtPath, encoding="utf-8") as f:
                next(f)
                lines = [line for line in f.readlines() if line.strip()]  # 移除空行
                aReturn = {}
                aReturn['type'] = ''
                aReturn['edited_data'] = []

                # 依檔名判別txt檔案
                if 'list_remove' in sTxtPath:
                    aReturn['type'] = 'remove'
                elif 'list_name' in sTxtPath:
                    aReturn['type'] = 'name_edit'
                elif 'list_addr' in sTxtPath:
                    aReturn['type'] = 'addr_edit'

                for line in lines:
                    if not len(line) or line.startswith('#'):  # 判斷是否為空行或註釋
                        continue
                    if 'list_name' in sTxtPath:
                        aTemp = []
                        line = line.split(' >>> ')

                        for j in ['\\', '"', '""', ' ']:  # 排除異常符號
                            if j in line[1]:
                                line[1] = line[1].replace(j, '')
                        aTemp.append(line[0].strip())
                        aTemp.append(line[1].strip())
                        if (len(line) == 3) and (line[2] != ''):
                            aTemp.append(line[2].strip())
                        else:
                            aTemp.append('')
                        aReturn['edited_data'].append(aTemp)

                    elif 'list_addr' in sTxtPath:
                        aTemp = []
                        line = line.split(' >>> ')
                        aTemp.append(line[0].strip())

                        for j in ['\\', '"', ' ', '臺']:  # 排除異常符號
                            if j == '臺':
                                if j in line[1]:
                                    line[1] = line[1].replace(j, '台')
                            else:
                                if j in line[1]:
                                    line[1] = line[1].replace(j, '')

                        aTemp.append(line[1].strip())
                        aReturn['edited_data'].append(aTemp)
                    elif 'list_remove' in sTxtPath:
                        aReturn['edited_data'].append(int(line.strip()))

                return aReturn
        except Exception as e:
            print("讀取txt檔案: %s\nline: %s" % (e, e.__traceback__.tb_lineno))


    def update_database_data(self, aDatas):
        """更新資料庫資料

        Args:
            aDatas (array)): 陣列資料
        """
        success_num = 0

        if aDatas['type'] != 'remove':
            for i in aDatas['edited_data']:

                aRes = self.oDB.select_single(
                    'lists_data', 'g_no = %s' % (i[0]))

                aFor_source = json.loads(aRes['for_main'])
                aFor_source['name'] = escape_string(aFor_source['name'].strip())

                if 'co_addr' in aFor_source.keys():
                    aFor_source['co_addr'] = escape_string(aFor_source['co_addr'].strip())

                if aDatas['type'] == 'name_edit':  # 更新店家名稱
                    aFor_source['correcting_name'] = escape_string(i[1].strip())
                    if i[2]:
                        aFor_source['subtitle'] = escape_string(i[2].strip())
                elif aDatas['type'] == 'addr_edit':  # 更新店家地址
                    if 'correcting_name' in aFor_source.keys():
                        aFor_source['correcting_name'] = escape_string(aFor_source['correcting_name'].strip())
                    if 'subtitle' in aFor_source.keys():
                        aFor_source['subtitle'] = escape_string(aFor_source['subtitle'].strip())
                    aFor_source['correcting_address'] = i[1]

                aFor_source['google_url'] = escape_string(aFor_source['google_url'].strip())

                sFor_source = json.dumps(aFor_source, ensure_ascii=False)

                # 更新防錯
                check_update_status = self.cover_data(i[0], sFor_source)
                if check_update_status:
                    success_num += 1
                else:
                    print(i[0])
                    self.cover_data(i[0], aRes['for_main'])

            self.sSuccess_log = "%s 成功修改數量: %s" % (aDatas['type'], success_num)
        else:
            for i in aDatas['edited_data']:
                self.oDB.update('lists_data', 'status = 9, modified = NOW()', 'g_no = %s' % (i))
                success_num += 1
            self.sRemove_log = "%s 移除數量: %s" % (aDatas['type'], success_num)


    def cover_data(self, g_no, sFor_source):
        """更新資料

        Args:
            g_no (int): 資料流水號
            sFor_source (str): JSON Format string
        Returns:
            bool: True or False
        """
        try:
            self.oDB.update('lists_data', 'for_main = NULL, modified = NOW()', 'g_no = %s' % (g_no))
            self.oDB.update('lists_data', "for_main = '%s', modified = NOW()" % (sFor_source), 'g_no = %s' % (g_no))
            return True
        except Exception as e:
            print("cover_data: %s" % (e))
        return False


    def cover_excel_file(self, sExcelFilePath, aDatas):
        """回存excel資料
        """
        # 讀取產業名單列表
        wb = load_workbook(sExcelFilePath)

        # 迴圈跑工作表
        for sheetname in wb.sheetnames:
            if sheetname == 'Sheet' or sheetname == "已審核資料":
                continue

            ws = wb[sheetname]
            rows = ws.rows

            # 判斷執行動作 (更新名稱/地址資料 or 更新啟用狀態)
            if aDatas['type'] != 'remove':
                for row in rows:

                    # 更新審核狀態
                    if (row[4].value == None) or (row[4].value == ''):
                        row[4].value = 'o'

                    for edited_data in aDatas['edited_data']:
                        if row[0].value == int(edited_data[0]):
                            # 判斷修改項目
                            if aDatas['type'] == 'name_edit':
                                row[1].value = edited_data[1]
                            elif aDatas['type'] == 'addr_edit':
                                row[2].value = edited_data[1]
            else:
                for row in rows:
                    # 更新審核狀態
                    if (row[4].value == None) or (row[4].value == ''):
                        row[4].value = 'o'
                    if row[0].value in aDatas['edited_data']:
                        row[3].value = ""

        wb.save(sExcelFilePath)


if __name__ == '__main__':
    Exc()
