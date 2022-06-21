#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import base
import openpyxl
import re


class Exc:
    """
        產生評分資料店家清單
    """
    # 資料庫連線
    oDB = base.DB()
    oDB.connect("twd_SHU")

    sFile_path = ''


    def __init__(self, sProject_no):

        self.assign_project_data(sProject_no)


    def assign_project_data(self, sProject_no):
        """產生指定專案excel資料
        Args:
            sProject_no (str): 專案流水號
        """
        try:
            aProject_data = self.oDB.select_single(
                "lists_data_project", "project_no = %s" % (sProject_no))

            aDatas = self.get_data(aProject_data['project_no'])

            aRes = self.addr_match_compare_process(aDatas)

            self.excel_file_production(aRes, aProject_data['project_name'])
        except Exception as e:
            print("產生指定專案excel資料: %s " % e)


    def get_data(self, sProject_no):
        """取得匯入主要lists_data資料
        Args:
            sProject_no (str): project_no
        Returns:
            [array]: 符合project_no陣列資料
        """
        try:
            sSql = "SELECT g_no, for_main->>'$.name' AS name, for_main->>'$.google_url' AS google_url, for_main->>'$.co_addr' AS co_addr, for_main->>'$.correcting_name' AS c_name, for_main->>'$.correcting_address' AS c_co_addr, status FROM lists_data WHERE project_no = %s" % (sProject_no)

            self.oDB.cursor.execute(sSql)
            self.oDB.db.commit()
            return self.oDB.cursor.fetchall()
        except Exception as e:
            print("取得匯入主要lists_data資料: %s" % e)


    def addr_match_compare_process(self, datas):
        """依照地址狀態分為'無地址資料'、'待確認地址資料'、'地址不需處理資料'、'已審核資料'四份，並做為excel工作表分類使用
        Args:
            datas (array): 指定專案的lists_data資料
        Returns:
            [array]: g_no, 名稱, 地址及啟用狀態陣列資料
        """
        try:

            # 符合條件資料資料
            Total_data = []
            # 待確認地址資料
            Total_data_check = []
            # 無地址資料
            Total_data_none = []
            # 已經審核資料
            Already_checked_data = []

            # 判斷無地址資料
            for data in datas:
                tmp_list = []
                tmp_list.append(data['g_no'])

                tmp_list.append(data['c_name']) if (
                    data['c_name']) else tmp_list.append(data['name'])

                tmp_list.append(data['c_co_addr']) if (
                    data['c_co_addr']) else tmp_list.append(data['co_addr'])

                tmp_list.append('v') if (
                    data['status'] != 9) else tmp_list.append('')

                tmp_list.append('o') if ((data['status'] == 2) or (
                    data['status'] == 3) or (data['status'] == 9)) else tmp_list.append('')

                tmp_list.append(data['google_url'])

                if data['status'] == 2 or data['status'] == 3 or data['status'] == 9:
                    Already_checked_data.append(tmp_list)

                elif data['c_co_addr'] is None:   # 無修正地址資料
                    if data['co_addr'] is None and (tmp_list not in Total_data_none):
                        Total_data_none.append(tmp_list)

                    elif['co_addr']:
                        match = re.search(
                            r'^(.[\u4e00-\u9fa5]?縣|.[\u4e00-\u9fa5]?市)(\D*?區|\D*?市|\D*?鄉|\D*?鎮)(\D*?路|\D*?街|\D*?大道)', data['co_addr'])

                        # 符合正則判斷
                        if match and (tmp_list not in Total_data):
                            Total_data.append(tmp_list)
                        # 待確認資料
                        if (match is None) and (tmp_list not in Total_data_check):
                            Total_data_check.append(tmp_list)
                else:
                    Total_data.append(tmp_list)

                self.oDB.update('lists_data', 'status = 1, modified = NOW()', 'g_no = %s AND status = 0' % (data['g_no']))

            return Total_data, Total_data_none, Total_data_check, Already_checked_data
        except Exception as e:
            print("解析資料地址: %s" % e)


    def excel_file_production(self, row_datas, sProject_name):
        """產生excle檔案,供企劃人工審合用
        Args:
            row_datas (array): 陣列資料
            sProject_name (str): 專案名稱
        """
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            Titles = ("編號", "店家名稱", "地址", "啟用", "審核")

            # 逐行寫入, 並依照判斷寫入符合判斷知工作表
            d_num = 2   # 起始行數
            rec_1 = 0
            rec_2 = 0
            rec_3 = 0
            rec_4 = 0

            # 已審核資料
            sheet_already_check = wb.create_sheet("已審核資料", 0)
            sheet_already_check.append(Titles)
            sheet_already_check.column_dimensions["B"].width = 80
            sheet_already_check.column_dimensions["C"].width = 100
            sheet_already_check.column_dimensions["D"].width = 10
            sheet_already_check.column_dimensions["E"].width = 10
            for data in row_datas[3]:
                sheet_already_check.append(data)
                sheet_already_check["B%s" % (d_num)].hyperlink = data[5]
                d_num += 1
                rec_4 += 1
            sheet_already_check.delete_cols(6)

            # Google_List 工作表
            sheet_Google_List = wb.create_sheet("Google_List", 0)
            sheet_Google_List.append(Titles)
            sheet_Google_List.column_dimensions["B"].width = 80
            sheet_Google_List.column_dimensions["C"].width = 100
            sheet_Google_List.column_dimensions["D"].width = 10
            sheet_Google_List.column_dimensions["E"].width = 10
            d_num = 2
            for data in row_datas[0]:
                sheet_Google_List.append(data)
                sheet_Google_List["B%s" % (d_num)].hyperlink = data[5]
                d_num += 1
                rec_1 += 1
            sheet_Google_List.delete_cols(6)

            # Google_List無地址 工作表
            sheet_no_addr = wb.create_sheet("Google_List無地址", 0)
            sheet_no_addr.append(Titles)
            sheet_no_addr.column_dimensions["B"].width = 80
            sheet_no_addr.column_dimensions["C"].width = 100
            sheet_no_addr.column_dimensions["D"].width = 10
            sheet_no_addr.column_dimensions["E"].width = 10
            d_num = 2
            for data in row_datas[1]:
                sheet_no_addr.append(data)
                sheet_no_addr["B%s" % (d_num)].hyperlink = data[5]
                d_num += 1
                rec_2 += 1
            sheet_no_addr.delete_cols(6)

            # Google_List待確認 工作表
            sheet_need_check = wb.create_sheet("Google_List待確認", 0)
            sheet_need_check.append(Titles)
            sheet_need_check.column_dimensions["B"].width = 80
            sheet_need_check.column_dimensions["C"].width = 100
            sheet_need_check.column_dimensions["D"].width = 10
            sheet_need_check.column_dimensions["E"].width = 10
            d_num = 2
            for data in row_datas[2]:
                sheet_need_check.append(data)
                sheet_need_check["B%s" % (d_num)].hyperlink = data[5]
                d_num += 1
                rec_3 += 1
            sheet_need_check.delete_cols(6)

            self.sFile_path = r"D:\\test\\Flask_test\\flask_sql\\各產業資料\\%s\\Google評分_%s_店家名單資訊.xlsx" % (sProject_name, sProject_name)
            wb.save(self.sFile_path)
        except Exception as e:
            print("產生excel檔案: %s" % e)


if __name__ == '__main__':
    oExc = Exc()
